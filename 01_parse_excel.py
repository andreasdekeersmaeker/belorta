"""
Step 1 – Parse the bell pepper pivot Excel into a clean long-format DataFrame.

The source file is a pivot table with a 3-level row hierarchy:
  Category code  (e.g. 198, 200, 201, 202, 203)
  Category name  (e.g. "198 - PAPRIKA DIV.")
  Vendor name    (e.g. "P  17 - GROBAR")

Columns repeat for every year (2013-2027) × quarter (Q1-Q4):
  Sum of Stuks   – quantity (pieces)
  Sum of Opp_Ha  – area in hectares (constant per vendor, repeated for every cell)

Output: bellpepper_long.csv  with columns:
  year | quarter | category_code | category_name | vendor | stuks | opp_ha
"""

import re
import pandas as pd
from openpyxl import load_workbook
from pathlib import Path

# ── CONFIG ──────────────────────────────────────────────────────────────────
EXCEL_PATH = Path("./bellpepers.xlsx")   # adjust if needed
OUTPUT_CSV = Path("bellpepper_long.csv")
# ────────────────────────────────────────────────────────────────────────────


def parse_pivot(excel_path: Path) -> pd.DataFrame:
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active

    all_rows = list(ws.iter_rows(values_only=True))
    header_year   = all_rows[1]   # row 2: year labels
    header_q      = all_rows[2]   # row 3: Q1-Q4
    header_type   = all_rows[3]   # row 4: Sum of Stuks / Sum of Opp_Ha
    data_rows     = all_rows[4:]  # row 5 onwards

    # Build a mapping: column index → (year, quarter, field)
    col_map = {}
    current_year = None
    current_q    = None

    for col_idx in range(1, len(header_year)):
        yr  = header_year[col_idx]
        q   = header_q[col_idx]
        typ = header_type[col_idx]

        if isinstance(yr, int) and 2000 < yr < 2030:
            current_year = yr
        if q in ("Q1", "Q2", "Q3", "Q4"):
            current_q = q

        if typ == "Sum of Stuks":
            col_map[col_idx] = (current_year, current_q, "stuks")
        elif typ == "Sum of Opp_Ha":
            col_map[col_idx] = (current_year, current_q, "opp_ha")

    # Parse data rows – detect hierarchy level from label format
    current_cat_code = None
    current_cat_name = None
    records = []

    cat_code_re   = re.compile(r"^\d{3}$")              # bare number: 198, 200 …
    cat_name_re   = re.compile(r"^\d{3}\s+-\s+")        # "198 - PAPRIKA DIV."
    vendor_re     = re.compile(r"^[PGp]\s*\d+\s+-\s+")  # "P  17 - GROBAR" or "G4903 - …"

    for row in data_rows:
        label = row[0]
        if label is None:
            continue

        label_str = str(label).strip()

        if cat_code_re.match(label_str):
            current_cat_code = int(label_str)
            continue  # subtotal row – skip
        elif cat_name_re.match(label_str):
            current_cat_name = label_str
            continue  # subtotal row – skip
        elif not vendor_re.match(label_str):
            continue  # unrecognised row (e.g. grand total)

        vendor = label_str

        # Collect all (year, quarter, stuks, opp_ha) from this row
        year_q_data: dict = {}
        for col_idx, (year, quarter, field) in col_map.items():
            if col_idx >= len(row):
                continue
            val = row[col_idx]
            key = (year, quarter)
            if key not in year_q_data:
                year_q_data[key] = {"stuks": None, "opp_ha": None}
            if val is not None:
                year_q_data[key][field] = float(val)

        for (year, quarter), vals in year_q_data.items():
            if year is None or quarter is None:
                continue
            records.append(
                {
                    "year":          year,
                    "quarter":       quarter,
                    "category_code": current_cat_code,
                    "category_name": current_cat_name,
                    "vendor":        vendor,
                    "stuks":         vals["stuks"],
                    "opp_ha":        vals["opp_ha"],
                }
            )

    df = pd.DataFrame(records)

    # opp_ha is constant per vendor – fill forward within vendor group
    df = df.sort_values(["vendor", "category_code", "year", "quarter"])
    df["opp_ha"] = df.groupby(["vendor", "category_code"])["opp_ha"].transform(
        lambda s: s.ffill().bfill()
    )

    # Remove rows that are entirely missing quantity AND area
    df = df.dropna(subset=["stuks", "opp_ha"], how="all").reset_index(drop=True)

    # Treat missing stuks as 0 (vendor didn't deliver that quarter)
    df["stuks"] = df["stuks"].fillna(0).astype(int)

    return df


if __name__ == "__main__":
    df = parse_pivot(EXCEL_PATH)
    df.to_csv(OUTPUT_CSV, index=False)
    print(f"Parsed {len(df):,} rows → {OUTPUT_CSV}")
    print(df.head(10).to_string())
    print("\nYears covered:", sorted(df["year"].unique()))
    print("Categories:", df["category_name"].unique())
    print("Vendors:", df["vendor"].nunique())
